import openpyxl
import random
import time
import os
# pip 호출

# def 함수 선언 단어 load (filepath)
def load_vocabulary(file_path): # 
    wb = openpyxl.load_workbook(file_path) # wb 라는 변수는 file_path에 있는 파일을 openpyxl.load_workbook 명렁
    sheet = wb.active # sheet 
    vocabulary = [(sheet.cell(row=row, column=1).value, sheet.cell(row=row, column=2).value, sheet.cell(row=row, column=3).value) for row in range(2, sheet.max_row + 1)]
    #voacbulary 라는 이름으로 sheet에 할당된 셀 1,2,3 생성
    return vocabulary # 함수 종료

def save_to_excel(data, file_path):# 데이터를 file_path 에 있는 파일에 저장
    
    if os.path.exists(file_path):  
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active # 존재하면 열고 
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet["A1"] = "단어"
        sheet["B1"] = "뜻"
        sheet["C1"] = "카운트" # abc에 저장

    for idx, (word, meaning, count) in enumerate(data, start=sheet.max_row + 1):#다음 행에 추가
        sheet.cell(row=idx, column=1).value = word
        sheet.cell(row=idx, column=2).value = meaning
        sheet.cell(row=idx, column=3).value = count  # 엑셀 파일에 카운트 값 저장
        

    wb.save(file_path)


def brain(vocabulary, file_path):
    random.seed(time.time())  # 현재 시간을 시드로 사용
    vocabulary.sort(key=lambda x: x[2])  # 카운트가 낮은 순서로 정렬
    current_index = 0  # 현재 인덱스 초기화
    
    while current_index < len(vocabulary):
        word, meaning, count = vocabulary[current_index]  # 현재 인덱스의 단어 선택
        
        print("-------------------------")
        print("||단어||:", word)
        print("-------------------------")  
        input("Enter")  # 사용자가 엔터를 입력하면 뜻이 출력됨
        
        print("-------------------------")
        print("||뜻||:", meaning)
        print("-------------------------")
        print("\n")
        print("\n")
        
        option = input("Enter 'y' for next word, 'p' for previous word, or type '!exit' to go back to mode selection: ")
        
        if option.lower() == 'y':
            # count + 1
            count += 1
            # 수정된 카운트를 원래의 vocabulary에 반영
            for idx, (w, m, c) in enumerate(vocabulary):
                if w == word:
                    vocabulary[idx] = (w, m, count)
                    
            current_index += 1  # 다음 단어로 인덱스 이동
            
        elif option.lower() == 'p':
            current_index = max(0, current_index - 1)  # 이전 단어로 인덱스 이동
            
        elif option.lower() == '!exit':
            print("Exiting brain mode...")
            # 현재 카운트 값을 엑셀 파일에 저장
            save_to_excel(vocabulary, file_path)
            break
        
        else:
            print("Invalid option! Please try again.")

    return vocabulary

def match(vocabulary):
    
    random.seed(time.time())
    
    while True:
        pair = random.choice(vocabulary)
        print("*****뜻을 맞추세요:", pair[0])
        answer = input("****[Enter!]: ")
        
    
        if answer.lower() == pair[1].lower():
            
            print("\n")
            print("-------------------------")
            print("****정답입니다!")
            print("-------------------------")
            print("\n")
            
        elif answer.lower() == "exitonnow":
            break
        
        else:
            print("\n")
            print("-------------------------")
            print("****오답입니다.")
            print("****뜻:", pair[1])  # 오답일 때 뜻 출력
            print("-------------------------")  
            print("\n")
        
        

def main():
    file_path = "C:\\english\\words_Ex.xlsx"  # 파일 경로 설정
    
    while True:
        mode = input("***   data, brain, match, !exit  *** ")
        
        if mode == "data":
            file_path = "C:\\english\\words_Ex.xlsx"
            data = []
            
            while True:
                print("-------------------------")
                word = input("***단어를 입력하세요 or exitonnow***    ")  # 공백을 제거하지 않음
                print("-------------------------")
                
                if word == "exitonnow":
                    break # 탈출 명령어
                
                if any(char.isalpha() and not char.isascii() for char in word):
                    print("영어 알파벳만 입력 가능합니다.")
                    continue # isalpha 로 유니코드 검사
                
                print("-------------------------")
                meaning = input("***뜻***       ")  
                print("-------------------------")
                
                if word.strip() == "" or meaning.strip() == "":
                    print("***null***")
                    continue #공백제거
                
                # 카운트를 1로 설정하여 데이터에 추가
                
                data.append((word, meaning, 1))  # 추가 시 카운트 1으로 초기화
                
            if data:
                save_to_excel(data, file_path)
                
                print("\n")
                print("********데이터가 저장되었습니다*******")
                print("\n")


        elif mode == "brain":
            # 암기 모드
            vocabulary = load_vocabulary(file_path)
            print("총", len(vocabulary), "개의 단어가 로드되었습니다.")
           
            input("암기 시작하려면 [Enter]")
            vocabulary = brain(vocabulary, file_path)  # file_path를 함께 전달
            save_to_excel(vocabulary, file_path)  # 엑셀 파일에 업데이트된 카운트 저장
            
        elif mode == "match":
            # 맞추기 모드
            vocabulary = load_vocabulary(file_path)
            print("총", len(vocabulary), "개의 단어가 로드되었습니다.")
            input("맞추기 퀴즈를 시작하려면 [Enter]를 누르세요.")
            match(vocabulary)
            
        elif mode == "!exit":
            print("프로그램을 종료합니다.")
            break
        else:
            print("올바른 모드를 선택하세요.")

if __name__ == "__main__":
    main()
