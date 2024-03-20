import openpyxl

def read_excel(file_path):
    """
    Read data from an Excel file.

    Parameters:
    - file_path (str): The path to the Excel file.

    Returns:
    - vocabulary (list): A list containing the data read from the Excel file.
    """
    # Open the Excel file
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Read data from the Excel file
    vocabulary = []
    for row in range(2, sheet.max_row + 1):
        word = sheet.cell(row=row, column=1).value
        meaning = sheet.cell(row=row, column=2).value
        count = sheet.cell(row=row, column=3).value
        
        # Check for None values
        if word is not None and meaning is not None and count is not None:
            vocabulary.append((word, meaning, count))
    
    # Sort data based on the first column
    vocabulary.sort(key=lambda x: x[0])
    
    return vocabulary

def remove_duplicates(file_path):
    """
    Remove duplicate rows from an Excel file.

    Parameters:
    - file_path (str): The path to the Excel file.

    Returns:
    - None
    """
    # Read data from the Excel file
    vocabulary = read_excel(file_path)
    
    # Remove duplicate rows
    unique_vocabulary = []
    for word, meaning, count in vocabulary:
        if (word, meaning, count) not in unique_vocabulary:
            unique_vocabulary.append((word, meaning, count))
    
    # Save unique data back to the Excel file
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['word', 'meaning', 'count'])
    for word, meaning, count in unique_vocabulary:
        sheet.append([word, meaning, count])
    
    # Save the updated Excel file
    wb.save(file_path)

# Example usage:
file_path = 'C:\\Words_PY\\words_EX.xlsx'
remove_duplicates(file_path)
